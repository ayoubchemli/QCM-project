import sys
import re
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from pathlib import Path
from typing import Dict, List

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import html
from datetime import datetime
import logging
from datetime import datetime, timedelta
import csv
import pandas as pd
import json
from quiz.views.quizesLevel import quizesLevel
from quiz.subject import Subject
from quiz.pdf_generator import generate_pdf



class ContactPage(QMainWindow):
    def __init__(self, parent=None, is_light_mode=False):
        super().__init__(parent)
        self.parent = parent
        self.setup_ui()
        self.apply_theme(is_light_mode)

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)  # Reduced from 30
        main_layout.setContentsMargins(30, 20, 30, 20)  # Reduced margins

        # Create main content card
        content_card = QFrame()
        content_card.setObjectName("contentCard")
        content_layout = QVBoxLayout(content_card)
        content_layout.setSpacing(20)  # Reduced from 30
        content_layout.setContentsMargins(20, 20, 20, 20)  # Reduced margins

        # Header with back button and title
        header_layout = QHBoxLayout()
        header_layout.setSpacing(10)  # Added spacing control
        
        back_button = HoverButton("← Return to Home")
        back_button.setFixedWidth(200)  
        back_button.clicked.connect(self.return_to_home)
        
        title = AnimatedLabel("Contact Us")
        title.setObjectName("contactTitle")
        
        header_layout.addWidget(back_button)
        header_layout.addWidget(title, alignment=Qt.AlignCenter)
        header_layout.addStretch()  # Better than fixed spacing
        
        content_layout.addLayout(header_layout)

        # Contact Methods Section
        methods_container = QFrame()
        methods_container.setObjectName("methodsContainer")
        methods_layout = QHBoxLayout(methods_container)
        methods_layout.setSpacing(15)  # Reduced from 20
        methods_layout.setContentsMargins(10, 10, 10, 10)  # Added margins

        contact_methods = [
            ("📧", "Email Us", "projet.cybersec.python@gmail.com", "Send us an email anytime"),
            ("📱", "Call Us", "+213 794 37 42 98", "Sun-Thu, 9:00-17:00"),
            ("💬", "Live Chat", "Available 24/7", "Chat with our support team")
        ]

        for icon, title, detail, description in contact_methods:
            card = self.create_contact_card(icon, title, detail, description)
            methods_layout.addWidget(card)

        content_layout.addWidget(methods_container)

        # Form and FAQ side by side
        bottom_container = QHBoxLayout()
        bottom_container.setSpacing(15)

        # Left side - Contact Form
        form_container = QFrame()
        form_container.setObjectName("formContainer")
        form_layout = QVBoxLayout(form_container)
        form_layout.setSpacing(15)  # Reduced from 20
        form_layout.setContentsMargins(15, 15, 15, 15)

        form_title = QLabel("Send us a Message")
        form_title.setObjectName("sectionTitle")
        form_layout.addWidget(form_title)

        # Form fields
        form_grid = QGridLayout()
        form_grid.setSpacing(10)  # Reduced from 15

        # Name fields in one row
        first_name = QLineEdit()
        first_name.setPlaceholderText("First Name")
        first_name.setObjectName("firstNameInput")  

        last_name = QLineEdit()
        last_name.setPlaceholderText("Last Name")
        last_name.setObjectName("lastNameInput")   
        
        form_grid.addWidget(first_name, 0, 0)
        form_grid.addWidget(last_name, 0, 1)

        # Other fields
        email = QLineEdit()
        email.setPlaceholderText("Email Address")
        email.setObjectName("emailInput")           # Changed from "formInput"
        form_grid.addWidget(email, 1, 0, 1, 2)

        subject = QLineEdit()
        subject.setPlaceholderText("Subject")
        subject.setObjectName("subjectInput")       # Changed from "formInput"
        form_grid.addWidget(subject, 2, 0, 1, 2)

        message = QTextEdit()
        message.setPlaceholderText("Your Message")
        message.setObjectName("messageInput")
        message.setMinimumHeight(100)  # Reduced from 150
        form_grid.addWidget(message, 3, 0, 1, 2)

        form_layout.addLayout(form_grid)

        submit_btn = HoverButton("Send Message")
        submit_btn.setObjectName("submitButton")
        submit_btn.setFixedWidth(150)  # Reduced from 200
        submit_btn.clicked.connect(self.submit_form)
        form_layout.addWidget(submit_btn, alignment=Qt.AlignCenter)

        # Right side - FAQ
        faq_container = QFrame()
        faq_container.setObjectName("faqContainer")
        faq_layout = QVBoxLayout(faq_container)
        faq_layout.setSpacing(10)
        faq_layout.setContentsMargins(15, 15, 15, 15)

        faq_title = QLabel("Frequently Asked Questions")
        faq_title.setObjectName("sectionTitle")
        faq_layout.addWidget(faq_title)

        faqs = [
            ("How do I reset my password?", 
            "Click on the 'Profile' button on the home page and fill your updated informations including password the 'Save Changes'."),
            ("Can I review my previous MCQ test attempts?", 
            "Yes! Visit the MCQ History page to see all your previous test scores and details."),
            ("How are the MCQ scores calculated?", 
            "Scores are calculated based on the number of correct answers. Each question carries equal marks."),
            ("Can I download old test results?", 
            "Ofc you can! Go to Export Results, choose your format and date range, then click Export.")
        ]

        for question, answer in faqs:
            faq_item = self.create_faq_item(question, answer)
            faq_layout.addWidget(faq_item)

        # Add form and FAQ to bottom container
        bottom_container.addWidget(form_container, 1)  # 1 is stretch factor
        bottom_container.addWidget(faq_container, 1)   # 1 is stretch factor

        content_layout.addLayout(bottom_container)
        main_layout.addWidget(content_card)

    def create_contact_card(self, icon, title, detail, description):
        card = QFrame()
        card.setObjectName("contactCard")
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        layout.setContentsMargins(15, 15, 15, 15)


        icon_label = QLabel(icon)
        icon_label.setObjectName("contactIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(icon_label)

        title_label = QLabel(title)
        title_label.setObjectName("contactTitle")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        detail_label = QLabel(detail)
        detail_label.setObjectName("contactDetail")
        detail_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(detail_label)

        desc_label = QLabel(description)
        desc_label.setObjectName("contactDescription")
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)

        return card

    def create_faq_item(self, question, answer):
        item = QFrame()
        item.setObjectName("faqItem")
        layout = QVBoxLayout(item)

        question_label = QLabel(question)
        question_label.setObjectName("faqQuestion")
        layout.addWidget(question_label)

        answer_label = QLabel(answer)
        answer_label.setObjectName("faqAnswer")
        answer_label.setWordWrap(True)
        layout.addWidget(answer_label)

        return item

    def validate_email(self, email):
        """Validate email format using regex pattern."""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None

    def sanitize_input(self, text):
        """Sanitize input to prevent XSS and other injection attacks."""
        return html.escape(text.strip())

    def create_email_template(self, first_name, last_name, email, subject, message_content):
        """Create a professional and visually appealing HTML email template."""
        current_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                /* Reset styles */
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                
                body {{
                    font-family: 'Segoe UI', Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    background-color: #f5f5f5;
                }}
                
                /* Container styles */
                .email-container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: #ffffff;
                    border-radius: 8px;
                    overflow: hidden;
                    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                }}
                
                /* Header styles */
                .header {{
                    background: linear-gradient(135deg, #4F46E5 0%, #818CF8 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                
                .header h2 {{
                    font-size: 24px;
                    margin-bottom: 10px;
                    font-weight: 600;
                }}
                
                .header p {{
                    font-size: 14px;
                    opacity: 0.9;
                }}
                
                /* Content styles */
                .content {{
                    padding: 30px;
                    background-color: #ffffff;
                }}
                
                .section {{
                    margin-bottom: 25px;
                    padding-bottom: 20px;
                    border-bottom: 1px solid #eee;
                }}
                
                .section:last-child {{
                    border-bottom: none;
                    margin-bottom: 0;
                    padding-bottom: 0;
                }}
                
                .section h3 {{
                    color: #4F46E5;
                    font-size: 18px;
                    margin-bottom: 15px;
                    font-weight: 600;
                }}
                
                .info-item {{
                    margin-bottom: 10px;
                }}
                
                .info-item strong {{
                    color: #1E293B;
                    font-weight: 600;
                }}
                
                .message-box {{
                    background-color: #f8fafc;
                    border-left: 4px solid #4F46E5;
                    padding: 20px;
                    border-radius: 4px;
                    margin-top: 15px;
                }}
                
                /* Footer styles */
                .footer {{
                    background-color: #1E293B;
                    color: #ffffff;
                    text-align: center;
                    padding: 20px;
                    font-size: 14px;
                }}
                
                .footer p {{
                    margin: 5px 0;
                    opacity: 0.9;
                }}
                
                .tag {{
                    display: inline-block;
                    background-color: #818CF8;
                    color: white;
                    padding: 3px 8px;
                    border-radius: 12px;
                    font-size: 12px;
                    margin-top: 10px;
                }}
                
                .social-links {{
                    margin-top: 15px;
                }}
                
                .social-links a {{
                    color: white;
                    text-decoration: none;
                    margin: 0 10px;
                }}
                
                /* Responsive design */
                @media only screen and (max-width: 600px) {{
                    .email-container {{
                        width: 100%;
                        border-radius: 0;
                    }}
                    
                    .header {{
                        padding: 20px;
                    }}
                    
                    .content {{
                        padding: 20px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <h2>📬 New Contact Form Submission</h2>
                    <p>Received on {current_date}</p>
                </div>
                
                <div class="content">
                    <div class="section">
                        <h3>👤 Contact Details</h3>
                        <div class="info-item">
                            <strong>Name:</strong> {first_name} {last_name}
                        </div>
                        <div class="info-item">
                            <strong>Email:</strong> {email}
                        </div>
                        <div class="info-item">
                            <strong>Subject:</strong> {subject}
                            <div class="tag">New Message</div>
                        </div>
                    </div>
                    
                    <div class="section">
                        <h3>💬 Message Content</h3>
                        <div class="message-box">
                            {message_content.replace(chr(10), '<br>')}
                        </div>
                    </div>
                </div>
                
                <div class="footer">
                    <p>This is an automated message from your contact form system</p>
                    <p>© 2024 AYOUB and ADAM. All rights reserved.</p>
                    <div class="social-links">
                        <a href="#">Discord</a> |
                        <a href="#">LinkedIn</a> |
                        <a href="#">Github</a>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        return html_content

    def submit_form(self):
        """Handle form submission with enhanced validation and error handling."""
        try:
            # Access the submit button and change its text
            submit_btn = self.findChild(QPushButton, "submitButton")
            original_text = submit_btn.text()
            submit_btn.setText("Sending...")
            submit_btn.setEnabled(False)

            # Gather and sanitize form inputs
            form_data = {
                'first_name': self.sanitize_input(self.findChild(QLineEdit, "firstNameInput").text()),
                'last_name': self.sanitize_input(self.findChild(QLineEdit, "lastNameInput").text()),
                'email': self.sanitize_input(self.findChild(QLineEdit, "emailInput").text()),
                'subject': self.sanitize_input(self.findChild(QLineEdit, "subjectInput").text()),
                'message': self.sanitize_input(self.findChild(QTextEdit, "messageInput").toPlainText())
            }

            # Validate all required fields
            for field, value in form_data.items():
                if not value:
                    raise ValueError(f"{field.replace('_', ' ').title()} is required")

            # Validate email format
            if not self.validate_email(form_data['email']):
                raise ValueError("Invalid email format")

            # Email configuration
            email_config = {
                'sender_email': "projet.cybersec.python@gmail.com",
                'sender_password': "anrc wogh zqfs xckj",  # Consider using environment variables
                'recipients_email': ["ayoubwork597@gmail.com", "adambelkadi1@gmail.com", "projet.cybersec.python@gmail.com"],
                'smtp_server': "smtp.gmail.com",
                'smtp_port': 587
            }

            # Create email message
            email_message = MIMEMultipart('alternative')
            email_message["From"] = email_config['sender_email']
            email_message["To"] = ", ".join(email_config['recipients_email'])
            email_message["Subject"] = f"Contact Form: {form_data['subject']}"
            
            # Create plain text and HTML versions
            text_content = f"""
            New Contact Form Submission

            Name: {form_data['first_name']} {form_data['last_name']}
            Email: {form_data['email']}
            Subject: {form_data['subject']}

            Message:
            {form_data['message']}
            """
            
            html_content = self.create_email_template(
                form_data['first_name'],
                form_data['last_name'],
                form_data['email'],
                form_data['subject'],
                form_data['message']
            )
            
            # Attach both versions
            email_message.attach(MIMEText(text_content, 'plain'))
            email_message.attach(MIMEText(html_content, 'html'))

            # Send email
            try:
                with smtplib.SMTP(email_config['smtp_server'], email_config['smtp_port']) as server:
                    server.starttls()
                    server.login(email_config['sender_email'], email_config['sender_password'])
                    # Explicitly pass the recipients list to send_message
                    server.sendmail(
                        email_config['sender_email'], 
                        email_config['recipients_email'],  # Recipients as a list
                        email_message.as_string()         # Full email as string
                    )
                # Log success
                logging.info(f"Email sent successfully to {', '.join(email_config['recipients_email'])}")
                # Show success message
                self.show_submit_success(submit_btn, original_text)
            except Exception as e:
                logging.error(f"Failed to send email: {e}")
            
            

        except ValueError as ve:
            error_message = str(ve)
            logging.warning(f"Validation error: {error_message}")
            self.show_error_message("Validation Error", error_message)
            submit_btn.setText(original_text)
            submit_btn.setEnabled(True)

        except smtplib.SMTPException as se:
            error_message = "Failed to send email. Please try again later."
            logging.error(f"SMTP error: {str(se)}")
            self.show_error_message("Email Error", error_message)
            submit_btn.setText(original_text)
            submit_btn.setEnabled(True)

        except Exception as e:
            error_message = "An unexpected error occurred. Please try again."
            logging.error(f"Unexpected error: {str(e)}")
            self.show_error_message("Error", error_message)
            submit_btn.setText(original_text)
            submit_btn.setEnabled(True)

    def show_error_message(self, title, message):
        """Display error message dialog."""
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setIcon(QMessageBox.Warning)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #1E293B;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #EF4444;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #DC2626;
            }
        """)
        msg.exec_()

    def show_submit_success(self, button, original_text):
        msg = QMessageBox(self)
        msg.setWindowTitle("Success")
        msg.setText("Message sent successfully!")
        msg.setInformativeText("We'll get back to you soon.")
        msg.setIcon(QMessageBox.Information)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #1E293B;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #4F46E5;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #4338CA;
            }
        """)
        
        button.setText(original_text)
        button.setEnabled(True)
        
        msg.exec_()

    def apply_theme(self, is_light_mode):
        if is_light_mode:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #F8FAFC, stop:1 #F1F5F9);
                }
                #contentCard {
                    background-color: white;
                    border-radius: 20px;
                    border: 1px solid #E2E8F0;
                }
                #contactCard {
                    background-color: #F8FAFC;
                    border: 1px solid #E2E8F0;
                    border-radius: 15px;
                    padding: 25px;
                }
                #contactIcon {
                    font-size: 36px;
                    margin-bottom: 10px;
                }
                #contactTitle {
                    color: #1E293B;
                    font-size: 20px;
                    font-weight: bold;
                }
                #contactDetail {
                    color: #4F46E5;
                    font-size: 16px;
                }
                #contactDescription {
                    color: #64748B;
                    font-size: 14px;
                }
                #formContainer, #faqContainer {
                    background-color: #F8FAFC;
                    border: 1px solid #E2E8F0;
                    border-radius: 15px;
                    padding: 25px;
                }
                #sectionTitle {
                    color: #1E293B;
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 20px;
                }
                #formInput, #firstNameInput, #lastNameInput, #emailInput, #subjectInput, #messageInput {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 8px;
                    padding: 12px;
                    color: #1E293B;
                }
                #formInput:focus, #firstNameInput:focus, #lastNameInput:focus, #emailInput:focus, #subjectInput:focus, #messageInput:focus {
                    border-color: #4F46E5;
                }
                #faqItem {
                    background-color: white;
                    border: 1px solid #E2E8F0;
                    border-radius: 10px;
                    padding: 15px;
                    margin-bottom: 10px;
                }
                #faqQuestion {
                    color: #1E293B;
                    font-size: 16px;
                    font-weight: bold;
                }
                #faqAnswer {
                    color: #64748B;
                    font-size: 14px;
                    margin-top: 8px;
                }
            """)
        else:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #0F172A, stop:1 #1E293B);
                }
                #contentCard {
                    background-color: #1E293B;
                    border-radius: 20px;
                    border: 1px solid #2D3748;
                }
                #contactCard {
                    background-color: #0F172A;
                    border: 1px solid #2D3748;
                    border-radius: 15px;
                    padding: 25px;
                }
                #contactIcon {
                    font-size: 36px;
                    margin-bottom: 10px;
                }
                #contactTitle {
                    color: white;
                    font-size: 20px;
                    font-weight: bold;
                }
                #contactDetail {
                    color: #818CF8;
                    font-size: 16px;
                }
                #contactDescription {
                    color: #94A3B8;
                    font-size: 14px;
                }
                #formContainer, #faqContainer {
                    background-color: #0F172A;
                    border: 1px solid #2D3748;
                    border-radius: 15px;
                    padding: 25px;
                }
                #sectionTitle {
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 20px;
                }
                #formInput,#firstNameInput, #lastNameInput, #emailInput, #subjectInput, #messageInput {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 8px;
                    padding: 12px;
                    color: white;
                }
                #formInput:focus,#firstNameInput:focus, #lastNameInput:focus, #emailInput:focus, #subjectInput:focus, #messageInput:focus {
                    border-color: #4F46E5;
                }
                #faqItem {
                    background-color: #1E293B;
                    border: 1px solid #2D3748;
                    border-radius: 10px;
                    padding: 15px;
                    margin-bottom: 10px;
                }
                #faqQuestion {
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                }
                #faqAnswer {
                    color: #94A3B8;
                    font-size: 14px;
                    margin-top: 8px;
                }
            """)

    def return_to_home(self):
        self.close()
        if self.parent:
            self.parent.show()

class ExportResultsPage(QMainWindow):
    def __init__(self, appstate, parent=None, is_light_mode=False):
        super().__init__(parent)
        self.parent = parent
        self.appstate = appstate
        self.selected_date_range = "all"  # Default to all time
        self.selected_format = "csv"  # Default to CSV
        self.setup_ui()
        self.apply_theme(is_light_mode)
        
    def update_preview_data(self):
        user_scores = self.appstate.getUser().scores
        filtered_scores = self.filter_scores_by_date(user_scores)
        
        self.preview_table.setRowCount(len(filtered_scores))
        
        for row, score in enumerate(filtered_scores):
            # Date
            date_item = QTableWidgetItem(score["date"])
            
            # course
            course = score['subject']['course']
            course_item = QTableWidgetItem(course)
            # chapter
            chapter = score['subject']['chapter']
            chapter_item = QTableWidgetItem(chapter)
            
            # Score as percentage
            score_percentage = f"{score['points']}%"
            score_item = QTableWidgetItem(score_percentage)
            
            status = "Passed" if score["points"] >= 50 else "Failed"
            status_item = QTableWidgetItem(status)
            
            # Set items
            items = [date_item, course_item, chapter_item, score_item, status_item]
            for col, item in enumerate(items):
                item.setTextAlignment(Qt.AlignCenter)
                self.preview_table.setItem(row, col, item)
                
    def filter_scores_by_date(self, scores):
        if self.selected_date_range == "custom":
            start_date = self.from_date.date().toPyDate()
            end_date = self.to_date.date().toPyDate()
        else:
            end_date = datetime.now().date()
            if self.selected_date_range == "Last 7 Days":
                start_date = end_date - timedelta(days=7)
            elif self.selected_date_range == "Last 30 Days":
                start_date = end_date - timedelta(days=30)
            elif self.selected_date_range == "Last 3 Months":
                start_date = end_date - timedelta(days=90)
            else:  # All Time
                return scores
        
        filtered_scores = []
        for score in scores:
            score_date = datetime.strptime(score["date"], 
                                        "%Y-%m-%d %H:%M:%S").date()
            if start_date <= score_date <= end_date:
                filtered_scores.append(score)
        
        return filtered_scores

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)  # Reduced from 30
        main_layout.setContentsMargins(30, 20, 30, 20)  # Reduced margins

        # Create main content card
        content_card = QFrame()
        content_card.setObjectName("contentCard")
        content_layout = QVBoxLayout(content_card)
        content_layout.setSpacing(20)  # Reduced from 30
        content_layout.setContentsMargins(30, 20, 30, 20)  # Reduced margins

        # Header with back button and title in one row
        header_layout = QHBoxLayout()
        header_layout.setSpacing(15)
        
        back_button = HoverButton("← Return to Home")
        back_button.setFixedWidth(200)  # Reduced width
        back_button.clicked.connect(self.return_to_home)
        
        title = AnimatedLabel("Export Results")
        title.setObjectName("exportTitle")
        
        header_layout.addWidget(back_button)
        header_layout.addWidget(title, alignment=Qt.AlignCenter)
        header_layout.addSpacing(150)  # Balance the layout
        
        content_layout.addLayout(header_layout)

        # Create a horizontal layout for options and preview
        main_content_layout = QHBoxLayout()
        main_content_layout.setSpacing(20)

        # Left side - Export Options
        options_container = QFrame()
        options_container.setFixedWidth(400)  # Fixed width for options
        options_container.setObjectName("optionsContainer")
        options_layout = QVBoxLayout(options_container)
        options_layout.setSpacing(15)  # Reduced spacing
        options_layout.setContentsMargins(20, 20, 20, 20)

        # Date Range Selection
        date_group = QFrame()
        date_group.setObjectName("optionGroup")
        date_layout = QVBoxLayout(date_group)
        date_layout.setSpacing(10)
        
        date_title = QLabel("Select Date Range")
        date_title.setObjectName("optionTitle")
        date_layout.addWidget(date_title)
        
        # Add "Use Custom Dates" checkbox
        self.use_custom_dates_checkbox = QCheckBox("Use Custom Dates")
        self.use_custom_dates_checkbox.stateChanged.connect(self.toggle_custom_dates)
        date_layout.addWidget(self.use_custom_dates_checkbox)

        date_buttons_layout = QGridLayout()  # Changed to grid layout
        date_buttons_layout.setSpacing(10)
        date_ranges = ["Last 7 Days", "Last 30 Days", "Last 3 Months", "All Time"]
        
        self.date_buttons = []
        for i, date_range in enumerate(date_ranges):
            btn = QPushButton(date_range)
            btn.setObjectName("dateRangeButton")
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            if date_range == "All Time":
                btn.setChecked(True)
            btn.clicked.connect(lambda checked, r=date_range: self.set_date_range(r))
            date_buttons_layout.addWidget(btn, i//2, i%2)
            self.date_buttons.append(btn)
        
        date_layout.addLayout(date_buttons_layout)
        options_layout.addWidget(date_group)

        # Custom Range Selection (in a more compact layout)
        custom_range = QFrame()
        custom_range.setObjectName("optionGroup")
        custom_layout = QGridLayout(custom_range)
        custom_layout.setSpacing(10)
        
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        
        custom_layout.addWidget(QLabel("From:"), 0, 0)
        custom_layout.addWidget(self.from_date, 0, 1)
        custom_layout.addWidget(QLabel("To:"), 1, 0)
        custom_layout.addWidget(self.to_date, 1, 1)
        
        options_layout.addWidget(custom_range)

        # Export Format Selection
        format_group = QFrame()
        format_group.setObjectName("optionGroup")
        format_layout = QVBoxLayout(format_group)
        format_layout.setSpacing(10)
        
        format_title = QLabel("Select Export Format")
        format_title.setObjectName("optionTitle")
        format_layout.addWidget(format_title)

        format_buttons_layout = QGridLayout()  # Changed to grid layout
        format_buttons_layout.setSpacing(10)
        formats = [
            ("CSV File", "csv", "📊"),
            ("JSON file", "json", "📝"),
            ("Excel File", "xlsx", "📘"),
            ("PDF Document", "pdf", "📄")
        ]
        
        for i, (format_name, format_id, icon) in enumerate(formats):
            btn = QPushButton(f"{icon} {format_name}")
            btn.setObjectName("formatButton")
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            if format_id == "csv":
                btn.setChecked(True)
            btn.clicked.connect(lambda checked, f=format_id: self.set_format(f))
            format_buttons_layout.addWidget(btn, i//2, i%2)
        
        format_layout.addLayout(format_buttons_layout)
        options_layout.addWidget(format_group)

        # Data Selection with refresh button
        data_group = QFrame()
        data_group.setObjectName("optionGroup")
        data_layout = QVBoxLayout(data_group)
        data_layout.setSpacing(10)
        
        # refresh button
        data_group = QFrame()
        data_group.setObjectName("optionGroup")
        data_layout = QVBoxLayout(data_group)
        data_layout.setSpacing(10)
        
        # Create centered refresh button
        refresh_btn = QPushButton("🔄 Refresh Preview")
        refresh_btn.setObjectName("refreshButton")
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setFixedSize(300, 50)  # Make button bigger
        refresh_btn.clicked.connect(self.update_preview_data)
        
        # Add button to layout with center alignment
        data_layout.addWidget(refresh_btn, alignment=Qt.AlignCenter)
        options_layout.addWidget(data_group)

        # Right side - Preview Section
        preview_container = QFrame()
        preview_container.setObjectName("previewContainer")
        preview_layout = QVBoxLayout(preview_container)
        preview_layout.setSpacing(15)
        
        preview_header = QHBoxLayout()
        preview_title = QLabel("Preview")
        preview_title.setObjectName("previewTitle")
        preview_header.addWidget(preview_title)
        
        preview_layout.addLayout(preview_header)
        
        # Preview table
        self.preview_table = QTableWidget()
        # self.preview_table.setRowCount(5)
        self.preview_table.setColumnCount(5)
        self.preview_table.setHorizontalHeaderLabels([
            "Date", "Course", "Chapter", "Score", "Status"
        ])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Add sample preview data
        sample_data = self.appstate.getUser().mcq_history()
        
        self.preview_table.setRowCount(len(sample_data))
        for row, (date, course, chapter, score, status) in enumerate(sample_data):
            date_item = QTableWidgetItem(date)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(row, 0, date_item)
            course_item = QTableWidgetItem(course)
            course_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(row, 1, course_item)
            chapter_item = QTableWidgetItem(chapter)
            chapter_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(row, 2, chapter_item)
            score_item = QTableWidgetItem(f"{score}%")
            score_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(row, 3, score_item)
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            self.preview_table.setItem(row, 4, status_item)
        
        # for row, data in enumerate(sample_data):
        #     for col, value in enumerate(data):
        #         item = QTableWidgetItem(value)
        #         item.setTextAlignment(Qt.AlignCenter)
        #         self.preview_table.setItem(row, col, item)
        
        preview_layout.addWidget(self.preview_table)

        # Add options and preview to main content layout
        main_content_layout.addWidget(options_container)
        main_content_layout.addWidget(preview_container, stretch=1)
        content_layout.addLayout(main_content_layout)

        # Export Button at the bottom
        export_btn = HoverButton("Export Results")
        export_btn.setObjectName("exportButton")
        export_btn.setFixedWidth(200)
        export_btn.setFixedHeight(50)
        export_btn.clicked.connect(self.export_results)
        
        content_layout.addWidget(export_btn, alignment=Qt.AlignCenter)

        # Add the content card to main layout
        main_layout.addWidget(content_card)
        
    def toggle_custom_dates(self, state):
        if state == Qt.Checked:
            self.selected_date_range = "custom"
            for btn in self.date_buttons:
                btn.setChecked(False)
                btn.setEnabled(False)
            self.from_date.setEnabled(True)
            self.to_date.setEnabled(True)
        else:
            for btn in self.date_buttons:
                btn.setEnabled(True)
            self.set_date_range("All Time")  # or default range

    def set_date_range(self, range_value):
        self.selected_date_range = range_value
        # Update buttons state
        for btn in self.date_buttons:
            btn.setChecked(btn.text() == range_value)

        if self.use_custom_dates_checkbox.isChecked():
            return  # Custom dates are being used

        if range_value == "Last 7 Days":
            self.from_date.setDate(QDate.currentDate().addDays(-7))
            self.to_date.setDate(QDate.currentDate())
            self.from_date.setEnabled(False)
            self.to_date.setEnabled(False)
        elif range_value == "Last 30 Days":
            self.from_date.setDate(QDate.currentDate().addDays(-30))
            self.to_date.setDate(QDate.currentDate())
            self.from_date.setEnabled(False)
            self.to_date.setEnabled(False)
        elif range_value == "Last 3 Months":
            self.from_date.setDate(QDate.currentDate().addMonths(-3))
            self.to_date.setDate(QDate.currentDate())
            self.from_date.setEnabled(False)
            self.to_date.setEnabled(False)
        elif range_value == "All Time":
            self.from_date.setDate(QDate(2024, 12, 26))
            self.to_date.setDate(QDate.currentDate())
            self.from_date.setEnabled(False)
            self.to_date.setEnabled(False)

    def set_format(self, format_value):
        self.selected_format = format_value
        format_map = {
            "📊 CSV File": "csv",
            "📝 JSON file": "json", 
            "📘 Excel File": "xlsx",
            "📄 PDF Document": "pdf"
        }
        # Update buttons state
        for btn in self.findChildren(QPushButton, "formatButton"):
            btn.setChecked(format_map[btn.text()] == format_value)

    def export_results(self):
        export_btn = self.findChild(HoverButton, "exportButton")
        original_text = export_btn.text()
        export_btn.setText("Exporting...")
        export_btn.setEnabled(False)

        try:
            # Get filtered data
            scores = self.filter_scores_by_date(self.appstate.getUser().scores)
            
            # Create output directory if it doesn't exist
            current_dir = Path(__file__).parent.parent.parent
            output_dir = current_dir / 'output'
            output_dir.mkdir(exist_ok=True)
            
            # Create filename base using current date
            date_str = datetime.now().strftime('%Y-%m-%d')
            base_filename = output_dir / f"results_{date_str}"
            
            if self.selected_format == "csv":
                self.export_to_csv(scores, f"{base_filename}.csv")
            elif self.selected_format == "json":
                self.export_to_json(scores, f"{base_filename}.json")
            elif self.selected_format == "xlsx":
                self.export_to_excel(scores, f"{base_filename}.xlsx")
            elif self.selected_format == "pdf":
                # Get the current date range for PDF generation
                if self.selected_date_range == "custom":
                    start_date = datetime.combine(self.from_date.date().toPyDate(), datetime.min.time())
                    end_date = datetime.combine(self.to_date.date().toPyDate(), datetime.max.time())
                else:
                    end_date = datetime.now()
                    if self.selected_date_range == "Last 7 Days":
                        start_date = end_date - timedelta(days=7)
                    elif self.selected_date_range == "Last 30 Days":
                        start_date = end_date - timedelta(days=30)
                    elif self.selected_date_range == "Last 3 Months":
                        start_date = end_date - timedelta(days=90)
                    else:  # All Time
                        start_date = None
                        end_date = None
                
                generate_pdf(self.appstate.getUser(), start_date, end_date)
                    
            self.show_export_success(export_btn, original_text)
        except Exception as e:
            self.show_export_error(str(e))
            export_btn.setText(original_text)
            export_btn.setEnabled(True)

    def export_to_csv(self, scores: List[Dict], filename: str) -> None:
        with open(filename, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Date', 'Course', 'Chapter', 'Score', 'Status'])
            for score in scores:
                writer.writerow([
                    score['date'],
                    score['subject']['course'],
                    score['subject']['chapter'],
                    f"{score['points']}%",
                    'Passed' if score['points'] >= 50 else 'Failed'
                ])

    def export_to_json(self, scores: List[Dict], filename: str) -> None:
        data = []
        for score in scores:
            data.append({
                'Date': score['date'],
                'Course': score['subject']['course'],
                'Chapter': score['subject']['chapter'],
                'Score': f"{score['points']}%",
                'Status': 'Passed' if score['points'] >= 50 else 'Failed'
            })
        with open(filename, 'w') as file:
            json.dump(data, file, indent=4)
        
    def export_to_excel(self, scores, filename):
    # Create DataFrame
        data = []
        for score in scores:
            data.append([
                score['date'],
                score['subject']['course'],
                score['subject']['chapter'],
                score['points'],
                'Passed' if score['points'] >= 50 else 'Failed'
            ])
        df = pd.DataFrame(data, columns=['Date', 'Course', 'Chapter', 'Score', 'Status'])
        
        # Create Excel writer object
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        
        # Write the data to Excel
        df.to_excel(writer, sheet_name='Results', index=False, startrow=1)
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Define styles
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True, color='FFFFFF', size=12)
        header_style.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            bottom=Side(style='medium', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0')
        )
        
        data_style = NamedStyle(name='data_style')
        data_style.font = Font(size=11)
        data_style.alignment = Alignment(horizontal='center', vertical='center')
        data_style.border = Border(
            bottom=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0')
        )
        
        # Add title
        title = f"Learning Results Report - Generated on {pd.Timestamp.now().strftime('%Y-%m-%d')}"
        worksheet.merge_cells('A1:E1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply styles
        for col in range(1, 6):
            cell = worksheet.cell(row=2, column=col)
            cell.style = header_style
            
        # Apply data styles and conditional formatting for scores and status
        for row in range(3, len(data) + 3):
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                cell.style = data_style
                
                # Special formatting for Score column
                if col == 4:
                    cell.value = float(str(cell.value).replace('%', ''))
                    cell.number_format = '0.0"%"'
                    if cell.value >= 80:
                        cell.fill = PatternFill(start_color='4ADE80', end_color='4ADE80', fill_type='solid')
                    elif cell.value >= 50:
                        cell.fill = PatternFill(start_color='FDE047', end_color='FDE047', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                
                # Special formatting for Status column
                if col == 5:
                    if cell.value == 'Passed':
                        cell.font = Font(color='22C55E', bold=True)
                    else:
                        cell.font = Font(color='EF4444', bold=True)
        
        # Adjust column widths
        for col in range(1, 6):
            worksheet.column_dimensions[get_column_letter(col)].width = 20
        
        # Add summary statistics
        summary_row = len(data) + 4
        worksheet.cell(row=summary_row, column=1, value='Summary Statistics').font = Font(bold=True, size=12)
        worksheet.merge_cells(f'A{summary_row}:E{summary_row}')
        
        stats = [
            ('Average Score', f"{df['Score'].mean():.1f}%"),
            ('Highest Score', f"{df['Score'].max():.1f}%"),
            ('Lowest Score', f"{df['Score'].min():.1f}%"),
            ('Pass Rate', f"{(df['Status'] == 'Passed').mean() * 100:.1f}%"),
            ('Total Assessments', len(df))
        ]
        
        for i, (label, value) in enumerate(stats):
            row = summary_row + i + 1
            worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
            worksheet.cell(row=row, column=2, value=value)
        
        # Add charts
        # Score Distribution Chart
        chart_row = summary_row + len(stats) + 2
        bar_chart = BarChart()
        bar_chart.title = "Score Distribution by Course"
        bar_chart.y_axis.title = 'Average Score (%)'
        bar_chart.x_axis.title = 'Course'
        
        course_avg = df.groupby('Course')['Score'].mean()
        data_refs = Reference(worksheet, min_col=4, min_row=2, max_row=len(data) + 2)
        cats_refs = Reference(worksheet, min_col=2, min_row=2, max_row=len(data) + 2)
        
        bar_chart.add_data(data_refs)
        bar_chart.set_categories(cats_refs)
        worksheet.add_chart(bar_chart, f'G{chart_row}')
        
        # Add pass/fail pie chart
        pie_chart = PieChart()
        pie_chart.title = "Pass/Fail Distribution"
        status_counts = df['Status'].value_counts()
        
        # Save and close
        writer.close()
        
        return filename
        
    def export_to_pdf(self):
        # Get the current date range
        if self.selected_date_range == "custom":
            start_date = datetime.combine(self.from_date.date().toPyDate(), datetime.min.time())
            end_date = datetime.combine(self.to_date.date().toPyDate(), datetime.max.time())
        else:
            end_date = datetime.now()
            if self.selected_date_range == "Last 7 Days":
                start_date = end_date - timedelta(days=7)
            elif self.selected_date_range == "Last 30 Days":
                start_date = end_date - timedelta(days=30)
            elif self.selected_date_range == "Last 3 Months":
                start_date = end_date - timedelta(days=90)
            else:  # All Time
                start_date = None
                end_date = None

        # Call generate_pdf with the date range
        generate_pdf(self.appstate.getUser(), start_date, end_date)

    def show_export_success(self, button, original_text):
        # Create success message
        msg = QMessageBox(self)
        msg.setWindowTitle("Success")
        msg.setText("Results exported successfully!")
        msg.setInformativeText(f"File saved as: results_{QDate.currentDate().toString('yyyy-MM-dd')}.{self.selected_format}")
        msg.setIcon(QMessageBox.Information)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #1E293B;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #4F46E5;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #4338CA;
            }
        """)
        
        # Reset button state
        button.setText(original_text)
        button.setEnabled(True)
        
        msg.exec_()

    def apply_theme(self, is_light_mode):
        # [Previous theme code remains the same, add these new styles:]
        if is_light_mode:
            self.setStyleSheet("""
                /* ... [Previous light theme styles] ... */
                #optionsContainer {
                    background-color: #F8FAFC;
                    border-radius: 15px;
                    border: 1px solid #E2E8F0;
                }
                #optionGroup {
                    background-color: white;
                    border-radius: 10px;
                    border: 1px solid #E2E8F0;
                    padding: 20px;
                }
                #optionTitle {
                    color: #1E293B;
                    font-size: 18px;
                    font-weight: bold;
                    margin-bottom: 15px;
                }
                #dateRangeButton, #formatButton {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 8px;
                    padding: 10px 20px;
                    color: #4A5568;
                }
                #dateRangeButton:checked, #formatButton:checked {
                    background-color: #4F46E5;
                    color: white;
                    border-color: #4F46E5;
                }
                QCheckBox {
                    color: #4A5568;
                    font-size: 14px;
                    padding: 5px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                #exportButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4F46E5, stop:1 #7C3AED);
                    color: white;
                    border: none;
                    border-radius: 25px;
                    font-weight: bold;
                }
                #exportButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4338CA, stop:1 #6D28D9);
                }
                #refreshButton {
                    background-color: #F1F5F9;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    color: #4A5568;
                }
                #refreshButton:hover {
                    background-color: #E2E8F0;
                }
                
                #optionsContainer {
                    background-color: #F8FAFC;
                    border-radius: 15px;
                    border: 1px solid #E2E8F0;
                }
                #optionGroup {
                    background-color: white;
                    border-radius: 10px;
                    border: 1px solid #E2E8F0;
                    padding: 20px;
                }
                #optionTitle {
                    color: #1E293B;
                    font-size: 18px;
                    font-weight: bold;
                    margin-bottom: 15px;
                }
                #dateRangeButton, #formatButton {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 8px;
                    padding: 10px 20px;
                    color: #4A5568;
                }
                #dateRangeButton:checked, #formatButton:checked {
                    background-color: #4F46E5;
                    color: white;
                    border-color: #4F46E5;
                }
                QCheckBox {
                    color: #4A5568;
                    font-size: 14px;
                    padding: 5px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                #refreshButton {
                    background-color: #4F46E5;
                    border: none;
                    border-radius: 8px;
                    padding: 8px 15px;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                }
                #refreshButton:hover {
                    background-color: #4338CA;
                }
                #exportButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4F46E5, stop:1 #7C3AED);
                    color: white;
                    border: none;
                    border-radius: 25px;
                    font-weight: bold;
                }
                #exportButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4338CA, stop:1 #6D28D9);
                }
                /* New styles for consistent text colors */
                QLabel {
                    color: #1E293B;
                }
                #previewTitle {
                    color: #1E293B;
                    font-size: 18px;
                    font-weight: bold;
                }
                QTableWidget {
                    color: #1E293B;
                    background-color: white;
                    border: 1px solid #E2E8F0;
                }
                QTableWidget::item {
                    color: #1E293B;
                }
                QTableWidget QHeaderView::section {
                    background-color: #F1F5F9;
                    color: #1E293B;
                    border: 1px solid #E2E8F0;
                }
                QDateEdit {
                    color: #1E293B;
                    background-color: white;
                    border: 1px solid #E2E8F0;
                    border-radius: 4px;
                    padding: 4px;
                }
                QDateEdit::drop-down {
                    border: none;
                }
                QDateEdit::down-arrow {
                    image: none;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 6px solid #4A5568;
                    width: 0;
                    height: 0;
                    margin-right: 6px;
                }
                
            """)
        else:
            self.setStyleSheet("""
                /* ... [Previous dark theme styles] ... */
                #optionsContainer {
                    background-color: #0F172A;
                    border-radius: 15px;
                    border: 1px solid #2D3748;
                }
                #optionGroup {
                    background-color: #1E293B;
                    border-radius: 10px;
                    border: 1px solid #2D3748;
                    padding: 20px;
                }
                #optionTitle {
                    color: white;
                    font-size: 18px;
                    font-weight: bold;
                    margin-bottom: 15px;
                }
                #dateRangeButton, #formatButton {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 8px;
                    padding: 10px 20px;
                    color: #E2E8F0;
                }
                #dateRangeButton:checked, #formatButton:checked {
                    background-color: #4F46E5;
                    color: white;
                    border-color: #4F46E5;
                }
                QCheckBox {
                    color: #E2E8F0;
                    font-size: 14px;
                    padding: 5px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                #exportButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4F46E5, stop:1 #7C3AED);
                    color: white;
                    border: none;
                    border-radius: 25px;
                    font-weight: bold;
                }
                #exportButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4338CA, stop:1 #6D28D9);
                }                
                #refreshButton {
                    background-color: #2D3748;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    color: #E2E8F0;
                }
                #refreshButton:hover {
                    background-color: #374151;
                }
                #optionsContainer {
                    background-color: #0F172A;
                    border-radius: 15px;
                    border: 1px solid #2D3748;
                }
                #optionGroup {
                    background-color: #1E293B;
                    border-radius: 10px;
                    border: 1px solid #2D3748;
                    padding: 20px;
                }
                #optionTitle {
                    color: white;
                    font-size: 18px;
                    font-weight: bold;
                    margin-bottom: 15px;
                }
                #dateRangeButton, #formatButton {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 8px;
                    padding: 10px 20px;
                    color: #E2E8F0;
                }
                #dateRangeButton:checked, #formatButton:checked {
                    background-color: #4F46E5;
                    color: white;
                    border-color: #4F46E5;
                }
                QCheckBox {
                    color: #E2E8F0;
                    font-size: 14px;
                    padding: 5px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                #refreshButton {
                    background-color: #4F46E5;
                    border: none;
                    border-radius: 8px;
                    padding: 8px 15px;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                }
                #refreshButton:hover {
                    background-color: #4338CA;
                }
                #exportButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4F46E5, stop:1 #7C3AED);
                    color: white;
                    border: none;
                    border-radius: 25px;
                    font-weight: bold;
                }
                #exportButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4338CA, stop:1 #6D28D9);
                }
                QLabel {
                    color: #E2E8F0;
                }
                #previewTitle {
                    color: white;
                    font-size: 18px;
                    font-weight: bold;
                }
                QTableWidget {
                    background-color: #1E293B;
                    border: none;
                    border-radius: 10px;
                    gridline-color: #2D3748;
                }
                QTableWidget::item {
                    padding: 12px;
                    color: #E2E8F0;
                }
                QHeaderView::section {
                    background-color: #0F172A;
                    color: #94A3B8;
                    padding: 12px;
                    border: none;
                    font-weight: bold;
                }
                QTableWidget::item:selected {
                    background-color: #312E81;
                    color: white;
                }
                QDateEdit {
                    color: #E2E8F0;
                    background-color: #1E293B;
                    border: 1px solid #2D3748;
                    border-radius: 4px;
                    padding: 4px;
                }
                QDateEdit::drop-down {
                    border: none;
                }
                QDateEdit::down-arrow {
                    image: none;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 6px solid #E2E8F0;
                    width: 0;
                    height: 0;
                    margin-right: 6px;
                }
            """)

    def return_to_home(self):
        self.close()
        if self.parent:
            self.parent.show()
                    
                    
class MCQHistoryPage(QMainWindow):
    def __init__(self, appstate, parent=None, is_light_mode=False):
        super().__init__(parent)
        self.parent = parent
        self.appstate = appstate
        self.setup_ui()
        self.apply_theme(is_light_mode)

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(30)
        main_layout.setContentsMargins(50, 40, 50, 40)

        # Create main content card
        content_card = QFrame()
        content_card.setObjectName("contentCard")
        content_layout = QVBoxLayout(content_card)
        content_layout.setSpacing(30)
        content_layout.setContentsMargins(40, 40, 40, 40)

        # Header with back button and title
        header_layout = QHBoxLayout()
        
        # Back button
        back_button = HoverButton("← Return to Home")
        back_button.setFixedWidth(200)
        back_button.clicked.connect(self.return_to_home)
        header_layout.addWidget(back_button)
        
        # Title
        title = AnimatedLabel("MCQ History")
        title.setObjectName("historyTitle")
        header_layout.addWidget(title, alignment=Qt.AlignCenter)
        header_layout.addSpacing(200)  # Balance the layout
        
        content_layout.addLayout(header_layout)

        # Statistics Overview Cards
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(20)

        # Create stat cards
        stat_cards = [
            ("Total Tests Taken", f"{self.appstate.getUser().count_tests_taken()}", "📝"),
            ("Average Score", f"{self.appstate.getUser().calculate_average_score()}%", "📊"),
            ("Best Performance", f"{self.appstate.getUser().get_best_score()}%", "🏆"),
            ("Tests This Month", f"{self.appstate.getUser().count_tests_this_month()}", "📅")
        ]

        for title, value, icon in stat_cards:
            card = self.create_stat_card(title, value, icon)
            stats_layout.addWidget(card)

        content_layout.addLayout(stats_layout)

        # Recent Tests Section
        recent_tests_container = QFrame()
        recent_tests_container.setObjectName("recentTestsContainer")
        recent_tests_layout = QVBoxLayout(recent_tests_container)

        # Section title
        section_title = QLabel("Recent Tests")
        section_title.setObjectName("sectionTitle")
        recent_tests_layout.addWidget(section_title)

        # Test history table
        self.history_table = QTableWidget()
        self.history_table.setObjectName("historyTable")
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels([
            "Date", "Course", "Chapter", "Score", "Status"
        ])
        
        # Add sample data
        self.add_sample_data()
        
        # Style the table
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        recent_tests_layout.addWidget(self.history_table)
        content_layout.addWidget(recent_tests_container)

        # Add the content card to main layout
        main_layout.addWidget(content_card)

    def create_stat_card(self, title, value, icon):
        card = QFrame()
        card.setObjectName("statCard")
        layout = QVBoxLayout(card)
        layout.setSpacing(10)

        # Icon
        icon_label = QLabel(icon)
        icon_label.setObjectName("statIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(icon_label)

        # Value
        value_label = QLabel(value)
        value_label.setObjectName("statValue")
        value_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)

        # Title
        title_label = QLabel(title)
        title_label.setObjectName("statTitle")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        return card

    def add_sample_data(self):
        # Sample test data
        test_data = self.appstate.getUser().mcq_history()
        # test_data = [
        #     ("2024-12-27", "Algebra", "85%", "45 mins", "Passed"),
        #     ("2024-12-25", "Data Structures", "92%", "60 mins", "Passed"),
        #     ("2024-12-23", "File Systems", "78%", "30 mins", "Passed"),
        #     ("2024-12-20", "Linear Algebra", "65%", "45 mins", "Failed"),
        #     ("2024-12-18", "Algorithms", "88%", "50 mins", "Passed"),
        #     ("2024-12-15", "Graph Theory", "73%", "40 mins", "Passed"),
        #     ("2024-12-12", "Number Theory", "95%", "55 mins", "Passed"),
        #     ("2024-12-10", "Binary Trees", "82%", "35 mins", "Passed")
        # ]

        self.history_table.setRowCount(len(test_data))
        for row, (date, course, chapter, score, status) in enumerate(test_data):
            self.history_table.setItem(row, 0, QTableWidgetItem(date))
            self.history_table.setItem(row, 1, QTableWidgetItem(course))
            self.history_table.setItem(row, 2, QTableWidgetItem(chapter))
            self.history_table.setItem(row, 3, QTableWidgetItem(score))
            
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            if status == "Passed":
                # status_item.setForeground(QColor("#10B981"))  # Green
                status_item.setBackground(QColor("#10B981"))  # Green
            else:
                # status_item.setForeground(QColor("#EF4444"))  # Red
                status_item.setBackground(QColor("#EF4444"))  # Red
            self.history_table.setItem(row, 4, status_item)

    def apply_theme(self, is_light_mode):
        if is_light_mode:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #F8FAFC, stop:1 #F1F5F9);
                }
                #contentCard {
                    background-color: white;
                    border-radius: 20px;
                    border: 1px solid #E2E8F0;
                }
                #historyTitle {
                    color: #1E293B;
                    font-size: 36px;
                    font-weight: bold;
                }
                #statCard {
                    background-color: #F8FAFC;
                    border: 1px solid #E2E8F0;
                    border-radius: 15px;
                    padding: 20px;
                    min-width: 200px;
                }
                #statIcon {
                    font-size: 32px;
                }
                #statValue {
                    color: #1E293B;
                    font-size: 28px;
                    font-weight: bold;
                }
                #statTitle {
                    color: #64748B;
                    font-size: 14px;
                }
                #recentTestsContainer {
                    background-color: #F8FAFC;
                    border-radius: 15px;
                    border: 1px solid #E2E8F0;
                    padding: 20px;
                }
                #sectionTitle {
                    color: #1E293B;
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 20px;
                }
                QTableWidget {
                    background-color: white;
                    border: none;
                    border-radius: 10px;
                    gridline-color: #E2E8F0;
                }
                QTableWidget::item {
                    padding: 12px;
                    color: #1E293B;
                }
                QHeaderView::section {
                    background-color: #F1F5F9;
                    color: #64748B;
                    padding: 12px;
                    border: none;
                    font-weight: bold;
                }
                QTableWidget::item:selected {
                    background-color: #EDE9FE;
                    color: #1E293B;
                }
            """)
        else:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #0F172A, stop:1 #1E293B);
                }
                #contentCard {
                    background-color: #1E293B;
                    border-radius: 20px;
                    border: 1px solid #2D3748;
                }
                #historyTitle {
                    color: white;
                    font-size: 36px;
                    font-weight: bold;
                }
                #statCard {
                    background-color: #0F172A;
                    border: 1px solid #2D3748;
                    border-radius: 15px;
                    padding: 20px;
                    min-width: 200px;
                }
                #statIcon {
                    font-size: 32px;
                }
                #statValue {
                    color: white;
                    font-size: 28px;
                    font-weight: bold;
                }
                #statTitle {
                    color: #94A3B8;
                    font-size: 14px;
                }
                #recentTestsContainer {
                    background-color: #0F172A;
                    border-radius: 15px;
                    border: 1px solid #2D3748;
                    padding: 20px;
                }
                #sectionTitle {
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 20px;
                }
                QTableWidget {
                    background-color: #1E293B;
                    border: none;
                    border-radius: 10px;
                    gridline-color: #2D3748;
                }
                QTableWidget::item {
                    padding: 12px;
                    color: #E2E8F0;
                }
                QHeaderView::section {
                    background-color: #0F172A;
                    color: #94A3B8;
                    padding: 12px;
                    border: none;
                    font-weight: bold;
                }
                QTableWidget::item:selected {
                    background-color: #312E81;
                    color: white;
                }
            """)

    def return_to_home(self):
        self.close()
        if self.parent:
            self.parent.show()
            
            
class PasswordField(QFrame):
    def __init__(self, parent=None, is_light_mode=False):
        super().__init__(parent)
        self.is_light_mode = is_light_mode
        self.setup_ui()
        self.setup_connections()

    def setup_ui(self):
        self.layout = QVBoxLayout(self)
        self.layout.setSpacing(15)
        
        # Password input container
        self.input_container = QFrame()
        input_layout = QHBoxLayout(self.input_container)
        input_layout.setContentsMargins(0, 0, 0, 0)
        
        # Password input field
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Enter your new password")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setObjectName("passwordInput")
        self.password_input.setMinimumHeight(45)
        
        # Toggle password visibility button
        self.toggle_btn = QPushButton()
        self.toggle_btn.setText("👁")
        self.toggle_btn.setCheckable(True)
        self.toggle_btn.setObjectName("togglePassword")
        self.toggle_btn.setMinimumSize(45, 45)
        
        input_layout.addWidget(self.password_input)
        input_layout.addWidget(self.toggle_btn)
        
        # Password strength indicator
        self.strength_container = QFrame()
        strength_layout = QHBoxLayout(self.strength_container)
        strength_layout.setContentsMargins(0, 0, 0, 0)
        
        self.strength_bar = QProgressBar()
        self.strength_bar.setTextVisible(False)
        self.strength_bar.setFixedHeight(8)
        self.strength_bar.setRange(0, 100)
        self.strength_bar.setObjectName("strengthBar")
        
        self.strength_label = QLabel("Password Strength")
        self.strength_label.setObjectName("strengthLabel")
        
        strength_layout.addWidget(self.strength_bar)
        strength_layout.addWidget(self.strength_label)
        
        # Requirements list
        self.requirements_list = QFrame()
        req_layout = QVBoxLayout(self.requirements_list)
        req_layout.setContentsMargins(0, 10, 0, 0)
        req_layout.setSpacing(10)
        
        self.requirements = {
            'length': QLabel("• At least 8 characters"),
            'uppercase': QLabel("• At least one uppercase letter"),
            'lowercase': QLabel("• At least one lowercase letter"),
            'number': QLabel("• At least one number"),
            'special': QLabel("• At least one special character")
        }
        
        for req in self.requirements.values():
            req.setObjectName("requirement")
            req_layout.addWidget(req)
        
        # Add everything to main layout
        self.layout.addWidget(self.input_container)
        self.layout.addWidget(self.strength_container)
        self.layout.addWidget(self.requirements_list)
        
        self.apply_styles()

    def setup_connections(self):
        self.toggle_btn.clicked.connect(self.toggle_password_visibility)
        self.password_input.textChanged.connect(self.check_password_strength)

    def toggle_password_visibility(self):
        if self.toggle_btn.isChecked():
            self.password_input.setEchoMode(QLineEdit.Normal)
            self.toggle_btn.setText("👁‍🗨")
        else:
            self.password_input.setEchoMode(QLineEdit.Password)
            self.toggle_btn.setText("👁")

    def check_password_strength(self):
        password = self.password_input.text()
        score = 0
        
        # Check requirements
        has_length = len(password) >= 8
        has_upper = any(c.isupper() for c in password)
        has_lower = any(c.islower() for c in password)
        has_number = any(c.isdigit() for c in password)
        has_special = any(not c.isalnum() for c in password)
        
        # Update requirement labels
        self.requirements['length'].setProperty("met", has_length)
        self.requirements['uppercase'].setProperty("met", has_upper)
        self.requirements['lowercase'].setProperty("met", has_lower)
        self.requirements['number'].setProperty("met", has_number)
        self.requirements['special'].setProperty("met", has_special)
        
        # Calculate score
        if has_length: score += 20
        if has_upper: score += 20
        if has_lower: score += 20
        if has_number: score += 20
        if has_special: score += 20
        
        # Update requirement styles
        for req in self.requirements.values():
            if req.property("met"):
                req.setStyleSheet("color: #10B981; font-size: 14px;")  # Green
            else:
                req.setStyleSheet(f"color: {'#64748B' if self.is_light_mode else '#94A3B8'}; font-size: 14px;")

        
        # Update strength indicator
        self.strength_bar.setValue(score)
        
        # Update strength label
        if score <= 20:
            strength_text = "Weak"
            strength_color = "#EF4444"  # Red
        elif score <= 60:
            strength_text = "Medium"
            strength_color = "#F59E0B"  # Yellow
        elif score <= 80:
            strength_text = "Strong"
            strength_color = "#10B981"  # Green
        else:
            strength_text = "very Strong"
            strength_color = "#0fa900"  # DARK Green
            
        self.strength_label.setText(strength_text)
        self.strength_label.setStyleSheet(f"color: {strength_color}; font-size: 14px; font-weight: bold;")
        self.strength_bar.setStyleSheet(f"""
            QProgressBar {{
                border-radius: 4px;
                background-color: {'#F8FAFC' if self.is_light_mode else '#0F172A'};
            }}
            QProgressBar::chunk {{
                border-radius: 4px;
                background-color: {strength_color};
            }}
        """)

    def get_password(self):
        return self.password_input.text()

    def set_password(self, password):
        self.password_input.setText(password)

    def apply_styles(self):
        if self.is_light_mode:
            # Light theme styles
            self.setStyleSheet("""
                QFrame {
                    background: transparent;
                }
                
                #passwordInput {
                    padding: 12px 15px;
                    border: 2px solid #E2E8F0;
                    border-radius: 10px;
                    font-size: 16px;
                    background-color: white;
                    min-width: 300px;
                    color: #1E293B;
                }
                
                #passwordInput:focus {
                    border-color: #4F46E5;
                }
                
                #togglePassword {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 10px;
                    padding: 5px;
                    font-size: 20px;
                    margin-left: 10px;
                    color: #1E293B;
                }
                
                #togglePassword:hover {
                    background-color: #F8FAFC;
                    border-color: #4F46E5;
                }
                
                #strengthBar {
                    border-radius: 4px;
                    background-color: #F8FAFC;
                    min-width: 200px;
                    border: 1px solid #E2E8F0;
                    height: 8px;
                }
                
                #strengthLabel {
                    font-size: 14px;
                    font-weight: bold;
                    margin-left: 15px;
                    min-width: 100px;
                    color: #1E293B;
                }
                
                #requirement {
                    font-size: 14px;
                    padding: 3px 0;
                    color: #64748B;
                }
                
                #requirement[met="true"] {
                    color: #10B981;
                }
            """)
        else:
            # Dark theme styles
            self.setStyleSheet("""
                QFrame {
                    background: transparent;
                }
                
                #passwordInput {
                    padding: 12px 15px;
                    border: 2px solid #2D3748;
                    border-radius: 10px;
                    font-size: 16px;
                    background-color: #1E293B;
                    min-width: 300px;
                    color: white;
                }
                
                #passwordInput:focus {
                    border-color: #4F46E5;
                }
                
                #togglePassword {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 10px;
                    padding: 5px;
                    font-size: 20px;
                    margin-left: 10px;
                    color: white;
                }
                
                #togglePassword:hover {
                    background-color: #2D3748;
                    border-color: #4F46E5;
                }
                
                #strengthBar {
                    border-radius: 4px;
                    background-color: #0F172A;
                    min-width: 200px;
                    border: 1px solid #2D3748;
                    height: 8px;
                }
                
                #strengthLabel {
                    font-size: 14px;
                    font-weight: bold;
                    margin-left: 15px;
                    min-width: 100px;
                    color: #E2E8F0;
                }
                
                #requirement {
                    font-size: 14px;
                    padding: 3px 0;
                    color: #94A3B8;
                }
                
                #requirement[met="true"] {
                    color: #10B981;
                }
            """)


class ProfilePage(QMainWindow):
    def __init__(self, appstate, parent=None, is_light_mode=False):
        super().__init__(parent)
        self.appstate = appstate
        self.parent = parent
        self.setup_ui(is_light_mode)
        self.apply_theme(is_light_mode)

    def setup_ui(self, theme):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(30)
        main_layout.setContentsMargins(50, 40, 50, 40)

        # Create a card-like container for the entire content
        content_card = QFrame()
        content_card.setObjectName("contentCard")
        content_layout = QVBoxLayout(content_card)
        content_layout.setSpacing(5)
        content_layout.setContentsMargins(40, 10, 40, 10)

        # Header with back button
        header_layout = QHBoxLayout()
        back_button = HoverButton("← Return to Home")
        back_button.setFixedWidth(200)
        back_button.clicked.connect(self.return_to_home)
        header_layout.addWidget(back_button)
        header_layout.addStretch()
        content_layout.addLayout(header_layout)

        # Profile Section
        profile_section = QHBoxLayout()
        
        # Profile Info
        profile_info = QVBoxLayout()
        
        # Title with animation
        title = AnimatedLabel("Change Password")
        title.setObjectName("profileTitle")
        profile_info.addWidget(title)

        # Subtitle
        subtitle = QLabel("You can either change your email by typing it in, change your old password, or do both!")
        subtitle.setObjectName("profileSubtitle")
        profile_info.addWidget(subtitle)
        
        profile_section.addSpacing(40)
        profile_section.addLayout(profile_info, stretch=1)
        content_layout.addLayout(profile_section)

        # Create a form container
        form_container = QFrame()
        form_container.setObjectName("formContainer")
        form_layout = QVBoxLayout(form_container)
        form_layout.setSpacing(1)
        form_layout.setContentsMargins(30, 30, 30, 30)

        # Email field
        email_container = QFrame()
        email_container.setObjectName("fieldContainer")
        email_layout = QVBoxLayout(email_container)
        email_layout.setSpacing(8)

        email_label = QLabel("📧 Email Address")
        email_label.setObjectName("fieldLabel")
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText(self.appstate.getUser().email)
        self.email_input.setObjectName("inputField")

        email_layout.addWidget(email_label)
        email_layout.addWidget(self.email_input)
        form_layout.addWidget(email_container)

        # Old Password field
        old_pass_container = QFrame()
        old_pass_container.setObjectName("fieldContainer")
        old_pass_layout = QVBoxLayout(old_pass_container)
        old_pass_layout.setSpacing(8)

        old_pass_label = QLabel("🔒 Current Password")
        old_pass_label.setObjectName("fieldLabel")
        self.old_pass_input = QLineEdit()
        self.old_pass_input.setPlaceholderText("Enter your current password")
        self.old_pass_input.setEchoMode(QLineEdit.Password)
        self.old_pass_input.setObjectName("inputField")

        old_pass_layout.addWidget(old_pass_label)
        old_pass_layout.addWidget(self.old_pass_input)
        form_layout.addWidget(old_pass_container)

        # New Password field (using PasswordField class)
        new_pass_container = QFrame()
        new_pass_container.setObjectName("fieldContainer")
        new_pass_layout = QVBoxLayout(new_pass_container)
        new_pass_layout.setSpacing(8)

        new_pass_label = QLabel("🔒 New Password")
        new_pass_label.setObjectName("fieldLabel")
        self.new_pass_field = PasswordField(is_light_mode=theme)

        new_pass_layout.addWidget(new_pass_label)
        new_pass_layout.addWidget(self.new_pass_field)
        form_layout.addWidget(new_pass_container)

        # Confirm New Password field
        confirm_pass_container = QFrame()
        confirm_pass_container.setObjectName("fieldContainer")
        confirm_pass_layout = QVBoxLayout(confirm_pass_container)
        confirm_pass_layout.setSpacing(8)

        confirm_pass_label = QLabel("🔒 Confirm New Password")
        confirm_pass_label.setObjectName("fieldLabel")
        self.confirm_pass_input = QLineEdit()
        self.confirm_pass_input.setPlaceholderText("Confirm your new password")
        self.confirm_pass_input.setEchoMode(QLineEdit.Password)
        self.confirm_pass_input.setObjectName("inputField")

        confirm_pass_layout.addWidget(confirm_pass_label)
        confirm_pass_layout.addWidget(self.confirm_pass_input)
        form_layout.addWidget(confirm_pass_container)

        content_layout.addWidget(form_container)

        # Save button with animation
        save_button = HoverButton("Save Changes")
        save_button.setObjectName("saveButton")
        save_button.setFixedWidth(200)
        save_button.setFixedHeight(50)
        save_button.clicked.connect(self.save_changes)
        
        # Button container for center alignment
        button_container = QHBoxLayout()
        button_container.addStretch()
        button_container.addWidget(save_button)
        button_container.addStretch()
        content_layout.addLayout(button_container)

        # Add the content card to main layout
        main_layout.addWidget(content_card)

    def apply_theme(self, is_light_mode):
        if is_light_mode:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #F8FAFC, stop:1 #F1F5F9);
                }
                #contentCard {
                    background-color: white;
                    border-radius: 20px;
                    border: 1px solid #E2E8F0;
                }
                #profileTitle {
                    color: #1E293B;
                    font-size: 36px;
                    font-weight: bold;
                    margin-bottom: 10px;
                }
                #profileSubtitle {
                    color: #64748B;
                    font-size: 16px;
                    margin-bottom: 20px;
                }
                #formContainer {
                    background-color: #F8FAFC;
                    border-radius: 15px;
                    border: 1px solid #E2E8F0;
                }
                #fieldContainer {
                    background: transparent;
                    margin-bottom: 10px;
                }
                #fieldLabel {
                    color: #4A5568;
                    font-size: 14px;
                    font-weight: bold;
                }
                #inputField {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 10px;
                    padding: 12px 15px;
                    font-size: 14px;
                    color: #1E293B;
                }
                #inputField:focus {
                    border-color: #4F46E5;
                }

            """)
        else:
            self.setStyleSheet("""
                QMainWindow {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #0F172A, stop:1 #1E293B);
                }
                #contentCard {
                    background-color: #1E293B;
                    border-radius: 20px;
                    border: 1px solid #2D3748;
                }
                #profileTitle {
                    color: white;
                    font-size: 36px;
                    font-weight: bold;
                    margin-bottom: 10px;
                }
                #profileSubtitle {
                    color: #94A3B8;
                    font-size: 16px;
                    margin-bottom: 20px;
                }
                #formContainer {
                    background-color: #0F172A;
                    border-radius: 15px;
                    border: 1px solid #2D3748;
                }
                #fieldContainer {
                    background: transparent;
                    margin-bottom: 10px;
                }
                #fieldLabel {
                    color: #E2E8F0;
                    font-size: 14px;
                    font-weight: bold;
                }
                #inputField {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 10px;
                    padding: 12px 15px;
                    font-size: 14px;
                    color: white;
                }
                #inputField:focus {
                    border-color: #4F46E5;
                }
            """)

    def return_to_home(self):
        self.close()
        if self.parent:
            self.parent.show()

    def validate_fields(self, profile_data):
        email = self.email_input.text()
        old_password = self.old_pass_input.text()
        new_password = self.new_pass_field.get_password()
        confirm_password = self.confirm_pass_input.text()
        
        # Check if user wants to change email
        if email:
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                self.show_error("Invalid email format")
                return False
            print(self.appstate.getUser().email)
            self.appstate.getUser().change_email(email)
            print(self.appstate.getUser().email)
            print("Email changed")
            
            
        
        # Check if user wants to change password
        if old_password or new_password or confirm_password:
            # Validate all password fields are filled
            if not old_password:
                self.show_error("Please enter your current password")
                return False
                
            if not new_password:
                self.show_error("Please enter a new password")
                return False
                
            if not confirm_password:
                self.show_error("Please confirm your new password")
                return False
                
            # Validate password confirmation
            if new_password != confirm_password:
                self.show_error("New passwords do not match")
                return False
            
            has_upper = any(c.isupper() for c in new_password)
            has_lower = any(c.islower() for c in new_password)
            has_digit = any(c.isdigit() for c in new_password)
            has_special = any(not c.isalnum() for c in new_password)
            valid = len(new_password) >= 8 and has_digit and (has_upper or has_lower or has_special)
            if not valid:
                self.show_error("Password must be at least 8 characters and contain any of : uppercase, lowercase, number, and special character")
                return False
            password_changed = self.appstate.getUser().change_password(old_password, new_password)
            if not password_changed:
                self.show_error("Failed to change password. Please check your current password")
                return False
            
            
        
        # Ensure at least one change is being made
        if not email and not (old_password or new_password or confirm_password):
            self.show_error("Please fill in either email or password fields to make changes")
            return False
        
        return True

    def save_changes(self):
        # Add loading animation
        save_button = self.findChild(HoverButton, "saveButton")
        original_text = save_button.text()
        save_button.setText("Saving...")
        save_button.setEnabled(False)
        
        # Get all the field values
        profile_data = {
            'email': self.email_input.text(),
            'old_password': self.old_pass_input.text(),
            'new_password': self.new_pass_field.get_password(),
            'confirm_password': self.confirm_pass_input.text()
        }
        
        # Validate fields
        if not self.validate_fields(profile_data):
            save_button.setText(original_text)
            save_button.setEnabled(True)
            return
        
        try:
            # TODO: Implement your password change logic here
            # Simulate network delay
            QTimer.singleShot(1500, lambda: self.show_save_success(save_button, original_text))
            
        except Exception as e:
            self.show_error(f"Failed to save changes: {str(e)}")
            save_button.setText(original_text)
            save_button.setEnabled(True)

    

    def show_error(self, message):
        msg = QMessageBox(self)
        msg.setWindowTitle("Error")
        msg.setText(message)
        msg.setIcon(QMessageBox.Critical)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #1E293B;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #EF4444;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #DC2626;
            }
        """)
        msg.exec_()

    def show_save_success(self, button, original_text):
        # Create a custom success message box
        msg = QMessageBox(self)
        msg.setWindowTitle("Success")
        msg.setText("Profile changes saved successfully!")
        msg.setIcon(QMessageBox.Information)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #1E293B;
            }
            QMessageBox QLabel {
                color: white;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #4F46E5;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #4338CA;
            }
        """)
        
        # Reset button state
        button.setText(original_text)
        button.setEnabled(True)
        
        msg.exec_()
class CustomDropdownMenu(QMenu):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_menu()
        
    def setup_menu(self):
        self.setFont(QFont('Segoe UI', 11))
        self.update_theme(False)  # Default to dark theme
        self.add_default_items()

    def update_theme(self, is_light_mode):
        if is_light_mode:
            self.setStyleSheet("""
                QMenu {
                    background-color: white;
                    border: 2px solid #E2E8F0;
                    border-radius: 15px;
                    padding: 8px;
                    min-width: 200px;
                }
                QMenu::item {
                    padding: 12px 24px;
                    color: #2D3748;
                    border-radius: 8px;
                    margin: 2px 5px;
                }
                QMenu::item:selected {
                    background-color: #F7FAFC;
                    color: #2D3748;
                }
                QMenu::separator {
                    height: 2px;
                    background: #E2E8F0;
                    margin: 6px 15px;
                }
            """)
        else:
            self.setStyleSheet("""
                QMenu {
                    background-color: #1E293B;
                    border: 2px solid #2D3748;
                    border-radius: 15px;
                    padding: 8px;
                    min-width: 200px;
                }
                QMenu::item {
                    padding: 12px 24px;
                    color: #E0E0E0;
                    border-radius: 8px;
                    margin: 2px 5px;
                }
                QMenu::item:selected {
                    background-color: #2D3748;
                    color: #FFFFFF;
                }
                QMenu::separator {
                    height: 2px;
                    background: #2D3748;
                    margin: 6px 15px;
                }
            """)

        
    def add_default_items(self):
        actions = {
            "👤 Profile": "View Profile and History",
            "📊 MCQ History": "View Past Test Scores",
            "💾 Export Results": "Export Results to File",
            "📞 Contact": "Contact Us",
            "❌ Sign out": "Account Access"
        }

        for icon_text, tooltip in actions.items():
            if icon_text == "❌ Sign out":
                self.addSeparator()
            action = self.addAction(icon_text)
            action.setToolTip(tooltip)

class CustomDropdownButton(QPushButton):
    def __init__(self, parent=None):
        super().__init__("☰", parent)
        self.setup_button()
        self.setup_menu()
        
    def setup_button(self):
        self.setFixedSize(70, 65)
        self.setFont(QFont('Segoe UI', 16))
        self.update_theme(False)  # Default to dark theme

    def update_theme(self, is_light_mode):
        if is_light_mode:
            self.setStyleSheet("""
                QPushButton {
                    background-color: white;
                    color: #2D3748;
                    font-weight: bold;
                    border: 2px solid #E2E8F0;
                    border-radius: 25px;
                    padding: 5px;
                    margin: 10px;
                }
                QPushButton:hover {
                    background-color: #F7FAFC;
                    border-color: #CBD5E0;
                    color: #2D3748;
                    transition: all 0.3s ease;
                }
                QPushButton:pressed {
                    background-color: #EDF2F7;
                    border-color: #CBD5E0;
                    padding: 6px 4px 4px 6px;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #1E293B;
                    color: #E0E0E0;
                    font-weight: bold;
                    border: 2px solid #2D3748;
                    border-radius: 25px;
                    padding: 5px;
                    margin: 10px;
                }
                QPushButton:hover {
                    background-color: #2D3748;
                    border-color: #4A5568;
                    color: #FFFFFF;
                    transition: all 0.3s ease;
                }
                QPushButton:pressed {
                    background-color: #374151;
                    border-color: #4A5568;
                    padding: 6px 4px 4px 6px;
                }
            """)
        
    def setup_menu(self):
        self.menu = CustomDropdownMenu(self)
        self.setMenu(self.menu)
        
    def add_custom_action(self, text, tooltip="", callback=None):
        """Add a custom action to the menu"""
        action = self.menu.addAction(text)
        action.setToolTip(tooltip)
        if callback:
            action.triggered.connect(callback)
            
    def add_separator(self):
        """Add a separator to the menu"""
        self.menu.addSeparator()

class ScrollButton(QPushButton):
    def __init__(self, direction, parent=None):
        super().__init__(parent)
        self.direction = direction
        self.setFixedSize(45, 45)  # Slightly smaller for elegance
        self.setCursor(Qt.PointingHandCursor)
        
        # Set arrow symbols based on direction using more elegant unicode arrows
        self.setText('❮' if direction == 'left' else '❯')
        
        # Create animation for hover effect
        self._animation = QPropertyAnimation(self, b"geometry")
        self._animation.setDuration(150)
        
        self.setStyleSheet("""
            QPushButton {
                background-color: rgba(79, 70, 229, 0.15);
                border: 2px solid rgba(79, 70, 229, 0.8);
                border-radius: 22px;
                color: rgba(79, 70, 229, 1);
                font-size: 18px;
                font-weight: bold;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: rgba(79, 70, 229, 1);
                border: 2px solid rgba(79, 70, 229, 1);
                color: white;
            }
            QPushButton:pressed {
                background-color: rgba(67, 56, 202, 1);
                border: 2px solid rgba(67, 56, 202, 1);
                color: white;
            }
        """)

    def enterEvent(self, event):
        geo = self.geometry()
        self._animation.setStartValue(geo)
        self._animation.setEndValue(QRect(geo.x()-2, geo.y()-2, 
                                        geo.width()+4, geo.height()+4))
        self._animation.setEasingCurve(QEasingCurve.OutQuad)
        self._animation.start()

    def leaveEvent(self, event):
        geo = self.geometry()
        self._animation.setStartValue(geo)
        self._animation.setEndValue(QRect(geo.x()+2, geo.y()+2, 
                                        geo.width()-4, geo.height()-4))
        self._animation.setEasingCurve(QEasingCurve.OutQuad)
        self._animation.start()

class HorizontalScrollArea(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)
        
        # Create scroll area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # Create container for content
        self.container = QWidget()
        self.container_layout = QHBoxLayout(self.container)
        self.container_layout.setSpacing(30)
        self.scroll_area.setWidget(self.container)
        
        # Create scroll buttons
        self.left_button = ScrollButton('left')
        self.right_button = ScrollButton('right')
        
        # Add widgets to layout
        self.layout.addWidget(self.left_button)
        self.layout.addWidget(self.scroll_area)
        self.layout.addWidget(self.right_button)
        
        # Connect buttons to scroll functions
        self.left_button.clicked.connect(self.scroll_left)
        self.right_button.clicked.connect(self.scroll_right)
        
        # Animation
        self.scroll_animation = QPropertyAnimation(self.scroll_area.horizontalScrollBar(), b"value")
        self.scroll_animation.setDuration(300)
        self.scroll_animation.setEasingCurve(QEasingCurve.OutCubic)

    def scroll_left(self):
        new_value = max(0, self.scroll_area.horizontalScrollBar().value() - 400)
        self.animate_scroll(new_value)

    def scroll_right(self):
        new_value = min(
            self.scroll_area.horizontalScrollBar().maximum(),
            self.scroll_area.horizontalScrollBar().value() + 400
        )
        self.animate_scroll(new_value)

    def animate_scroll(self, new_value):
        self.scroll_animation.setStartValue(self.scroll_area.horizontalScrollBar().value())
        self.scroll_animation.setEndValue(new_value)
        self.scroll_animation.start()

    def add_widget(self, widget):
        self.container_layout.addWidget(widget)

class ThemeToggleButton(QPushButton):
   def __init__(self, parent=None):
       super().__init__(parent)
       self.setCheckable(True)
       self.setCursor(Qt.PointingHandCursor)
       self.setFixedSize(60, 30)
       self.light_icon = "🌞"
       self.dark_icon = "🌙"
       self.setText(self.dark_icon)
       self.update_style(False)
       
   def update_style(self, is_light_mode=False):
       self.setChecked(is_light_mode)
       self.setText(self.light_icon if is_light_mode else self.dark_icon)
       style = """
           QPushButton {
               background-color: %s;
               border: 2px solid #4F46E5;
               border-radius: 15px;
               color: %s;
               font-size: 16px;
           }
           QPushButton:checked {
               background-color: %s;
               border: 2px solid #4F46E5;
               color: %s;
           }
       """ % (('#E2E8F0' if is_light_mode else '#1E293B'),
              ('black' if is_light_mode else 'white'),
              ('#1E293B' if is_light_mode else '#E2E8F0'),
              ('white' if is_light_mode else 'black'))
       self.setStyleSheet(style)

class HoverButton(QPushButton):
    def __init__(self, text):
        super().__init__(text)
        self.setFixedHeight(45)
        self.setCursor(Qt.PointingHandCursor)
        self._animation = QPropertyAnimation(self, b"geometry")
        self._animation.setDuration(150)
        self._original_geometry = None
        self._animation.finished.connect(self._on_animation_finished)
        self.setStyleSheet("""
            HoverButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #4F46E5, stop:1 #7C3AED);
                color: white;
                border: none;
                border-radius: 22px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            HoverButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #4338CA, stop:1 #6D28D9);
            }
        """)

    def showEvent(self, event):
        super().showEvent(event)
        # Store the original geometry when button is first shown
        self._original_geometry = self.geometry()

    def enterEvent(self, event):
        if self._original_geometry:
            self._animation.setStartValue(self._original_geometry)
            expanded = QRect(
                self._original_geometry.x() - 2,
                self._original_geometry.y() - 2,
                self._original_geometry.width() + 4,
                self._original_geometry.height() + 4
            )
            self._animation.setEndValue(expanded)
            self._animation.setEasingCurve(QEasingCurve.OutQuad)
            self._animation.start()
        super().enterEvent(event)

    def leaveEvent(self, event):
        if self._original_geometry:
            self._animation.setStartValue(self.geometry())
            self._animation.setEndValue(self._original_geometry)
            self._animation.setEasingCurve(QEasingCurve.OutQuad)
            self._animation.start()
        super().leaveEvent(event)
        
    def _on_animation_finished(self):
        # If mouse is not over button, ensure we're at original geometry
        if not self.underMouse() and self._original_geometry:
            self.setGeometry(self._original_geometry)

class GradientCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("courseCard")
        
        # Add shadow effect
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 60))
        self.setGraphicsEffect(shadow)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        path = QPainterPath()
        path.addRoundedRect(self.rect(), 15, 15)
        
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0, QColor("#1E293B"))
        gradient.setColorAt(1, QColor("#2D3B4F"))
        
        painter.fillPath(path, gradient)

class AnimatedLabel(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self._animation = QPropertyAnimation(self, b"geometry")
        self._animation.setDuration(200)
        self._original_geometry = None
        self._animation.finished.connect(self._on_animation_finished)
        
    def showEvent(self, event):
        super().showEvent(event)
        # Store the original geometry when label is first shown
        self._original_geometry = self.geometry()
        
    def enterEvent(self, event):
        if self._original_geometry:
            self._animation.setStartValue(self._original_geometry)
            elevated = QRect(
                self._original_geometry.x(),
                self._original_geometry.y() - 2,
                self._original_geometry.width(),
                self._original_geometry.height()
            )
            self._animation.setEndValue(elevated)
            self._animation.start()
        super().enterEvent(event)

    def leaveEvent(self, event):
        if self._original_geometry:
            self._animation.setStartValue(self.geometry())
            self._animation.setEndValue(self._original_geometry)
            self._animation.start()
        super().leaveEvent(event)
        
    def _on_animation_finished(self):
        # If mouse is not over label, ensure we're at original geometry
        if not self.underMouse() and self._original_geometry:
            self.setGeometry(self._original_geometry)

class MCQHomePage(QMainWindow):
    
   def setup_menu_actions(self):
        # Connect menu actions to their respective functions
        for action in self.dropdown_button.menu.actions():
            if action.text() == "👤 Profile":
                action.triggered.connect(self.open_profile)
            elif action.text() == "📊 MCQ History":
                action.triggered.connect(self.open_mcq_history)
            elif action.text() == "💾 Export Results":
                action.triggered.connect(self.open_export_results)
            elif action.text() == "📞 Contact":
                action.triggered.connect(self.open_contact)
            elif action.text() == "❌ Sign out":
                action.triggered.connect(self.sign_out)
   def sign_out(self):       
        from quiz.views.loginRegister import MCQApp

        # Create the login page
        self.login_page = MCQApp(self.appstate)
        self.login_page.showFullScreen()

        # Set a delay before closing self
        QTimer.singleShot(200, self.close)  # 200 milliseconds delay

   def open_contact(self):
        self.contact_page = ContactPage(self, self.theme_toggle.isChecked())
        self.contact_page.showFullScreen()

   def open_export_results(self):
        self.export_results_page = ExportResultsPage(self.appstate, self, self.theme_toggle.isChecked())
        self.export_results_page.showFullScreen()
   def open_mcq_history(self):
        self.mcq_history_page = MCQHistoryPage(self.appstate, self, self.theme_toggle.isChecked())
        self.mcq_history_page.showFullScreen()

   def open_profile(self):
        self.profile_page = ProfilePage(self.appstate, self, self.theme_toggle.isChecked())
        self.profile_page.showFullScreen()
        
   def setup_themes(self):
       self.dark_theme = {
           "main_bg": "background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #0F172A, stop:1 #1E293B)",
           "card_bg": "#1E293B",
           "text_primary": "white",
           "text_secondary": "#94A3B8",
           "button_gradient": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4F46E5, stop:1 #7C3AED)",
           "button_hover": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4338CA, stop:1 #6D28D9)",
           "new_badge": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10B981, stop:1 #059669)"
       }
       self.light_theme = {
            "main_bg": "background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #FFFFFF, stop:1 #F0F4F8)",
            "card_bg": "#FFFFFF",
            "text_primary": "#2D3748",  # Darker text for better contrast
            "text_secondary": "#4A5568", # Warmer gray for secondary text
            "button_gradient": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4F46E5, stop:1 #7C3AED)", # Keeping the button gradient
            "button_hover": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4338CA, stop:1 #6D28D9)",
            "new_badge": "qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10B981, stop:1 #059669)"
        }

   def create_course_card(self, title, description, chapters, is_new=False):
    card = QFrame()
    card.setObjectName("courseCard")
    layout = QVBoxLayout(card)
    layout.setSpacing(15)

    header = QHBoxLayout()
    title_label = QLabel(title)
    title_label.setObjectName("title")
    # Remove the hardcoded color here - let the theme handle it
    title_label.setStyleSheet("font-size: 24px; font-weight: bold;")
    header.addWidget(title_label)

    if is_new:
        new_badge = QLabel("New")
        new_badge.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                stop:0 #10B981, stop:1 #059669);
            color: white;
            border-radius: 12px;
            padding: 6px 12px;
            font-size: 14px;
            font-weight: bold;
            max-width: 60px;
        """)
        header.addWidget(new_badge)
    header.addStretch()
    layout.addLayout(header)

    desc = QLabel(description)
    desc.setObjectName("desc")
    desc.setWordWrap(True)
    # Remove hardcoded color here too
    desc.setStyleSheet("font-size: 14px; line-height: 1.5;")
    layout.addWidget(desc)

    chapters_label = QLabel("Chapters:")
    # Remove hardcoded color
    chapters_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
    layout.addWidget(chapters_label)

    for chapter in chapters:
        ch_label = QLabel(f"• {chapter}")
        ch_label.setProperty("class", "chapter")  # Add a class property for styling
        # Remove hardcoded color
        ch_label.setStyleSheet("margin-left: 15px;")
        layout.addWidget(ch_label)

    layout.addStretch()
    enroll_btn = HoverButton("Enroll Now")
    enroll_btn.clicked.connect(lambda checked, t=title: self.handle_enroll(t))
    layout.addWidget(enroll_btn)

    shadow = QGraphicsDropShadowEffect()
    shadow.setBlurRadius(15)
    shadow.setColor(QColor(0, 0, 0, 80))
    shadow.setOffset(0, 4)
    card.setGraphicsEffect(shadow)

    # Remove hardcoded colors from card styling
    card.setStyleSheet("""
        QFrame#courseCard {
            border-radius: 20px;
            padding: 25px;
            min-height: 350px;
        }
    """)
    return card


        
   def handle_enroll(self, course_title):
    print(f"Enrolling in course: {course_title}")
    self.appstate.setCourse(course_title)

    # Print the number of widgets and their names before cleanup
    print(f"Before cleanup, widgets in stack: {self.stacked_widget.count()}")
    for index in range(self.stacked_widget.count()):
        widget = self.stacked_widget.widget(index)
        print(f"Index {index}: {widget}")

    # Iterate through all widgets in the QStackedWidget (excluding index 0)
    for index in range(self.stacked_widget.count() - 1, 0, -1):  # Start from the last index down to 1
        widget = self.stacked_widget.widget(index)
        if widget:
            # Remove and schedule the widget for deletion
            self.stacked_widget.removeWidget(widget)
            widget.deleteLater()

    # Print the number of widgets and their names after cleanup
    print(f"After cleanup, widgets in stack: {self.stacked_widget.count()}")
    for index in range(self.stacked_widget.count()):
        widget = self.stacked_widget.widget(index)
        print(f"Index {index}: {widget}")
    
    # Pass the current theme state to quizesLevel
    is_light_mode = self.theme_toggle.isChecked()
    self.quizeslevel = quizesLevel(self.appstate, is_light_mode)
    
    # # Add a reference to the theme toggle
    # self.quizeslevel.theme_toggle = self.theme_toggle
    
    # # Connect the theme toggle to quizesLevel's apply_theme method
    # self.quizeslevel.theme_toggle.clicked.connect(
    #     lambda: self.quizeslevel.apply_theme(self.theme_toggle.isChecked())
    # )
    
    self.stacked_widget.addWidget(self.quizeslevel)
    self.stacked_widget.setCurrentIndex(1)

   def apply_theme(self, is_light_mode=False):
    theme = self.light_theme if is_light_mode else self.dark_theme
    self.theme_toggle.update_style(is_light_mode)
    
    # Update dropdown menu and button themes
    self.dropdown_button.update_theme(is_light_mode)
    self.dropdown_button.menu.update_theme(is_light_mode)
    
    # Main window and scrollbar styling
    self.setStyleSheet(f"""
        QMainWindow {{
            {theme["main_bg"]};
        }}
        QScrollArea {{
            border: none;
            background-color: transparent;
        }}
        QScrollBar:horizontal {{
            height: 8px;
            background: rgba(0, 0, 0, 0.1);
            border-radius: 4px;
            margin: 0px;
        }}
        QScrollBar::handle:horizontal {{
            background: #4F46E5;
            border-radius: 4px;
            min-width: 40px;
        }}
        QScrollBar::handle:horizontal:hover {{
            background: #6D28D9;
        }}
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
            width: 0px;
        }}
    """)

    # Card styling based on theme
    if is_light_mode:
        card_style = """
            QFrame#courseCard {
                background-color: white;
                border-radius: 20px;
                padding: 25px;
                min-height: 350px;
                border: 1px solid #E2E8F0;
            }
            QFrame#courseCard:hover {
                background-color: #F8FAFC;
            }
            QFrame#courseCard QLabel#title {
                color: #2D3748;
                font-size: 24px;
                font-weight: bold;
            }
            QFrame#courseCard QLabel#desc {
                color: #4A5568;
                font-size: 14px;
                line-height: 1.5;
            }
            QFrame#courseCard QLabel {
                color: #4A5568;
            }
            QFrame#courseCard QLabel[class="chapter"] {
                color: #4A5568;
                margin-left: 15px;
            }
        """
    else:
        card_style = """
            QFrame#courseCard {
                background-color: #1E293B;
                border-radius: 20px;
                padding: 25px;
                min-height: 350px;
            }
            QFrame#courseCard:hover {
                background-color: #233043;
            }
            QFrame#courseCard QLabel#title {
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
            QFrame#courseCard QLabel#desc {
                color: #94A3B8;
                font-size: 14px;
                line-height: 1.5;
            }
            QFrame#courseCard QLabel {
                color: #94A3B8;
            }
            QFrame#courseCard QLabel[class="chapter"] {
                color: #94A3B8;
                margin-left: 15px;
            }
        """

    # Apply card styling to all course cards
    for card in self.findChildren(QFrame, "courseCard"):
        card.setStyleSheet(card_style)

    # Update title and subtitle styling
    title_style = f"color: {theme['text_primary']}; font-size: 48px; font-weight: bold; margin-bottom: 20px;"
    subtitle_style = f"color: {theme['text_secondary']}; font-size: 18px; margin-bottom: 30px;"
    
    for label in self.findChildren(QLabel):
        if label.text().startswith("welcome"):
            label.setStyleSheet(title_style)
        elif label.text().startswith("Glad to see you"):
            label.setStyleSheet(subtitle_style)

   def __init__(self, appstate):
        super().__init__()
        self.appstate = appstate

        
        self.central_widget = QWidget()
        self.stacked_widget = QStackedWidget()
        self.setCentralWidget(self.stacked_widget)
        self.stacked_widget.addWidget(self.central_widget)
        main_layout = QVBoxLayout(self.central_widget)
        main_layout.setSpacing(40)
        
        # Create header container
        header_container = QWidget()
        header_layout = QHBoxLayout(header_container)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(0)

        # Add dropdown button to the left
        self.dropdown_button = CustomDropdownButton()
        header_layout.addWidget(self.dropdown_button, alignment=Qt.AlignLeft | Qt.AlignTop)

        
        self.setup_menu_actions()

        # Create center content container
        center_content = QWidget()
        center_layout = QVBoxLayout(center_content)
        center_layout.setSpacing(20)

        title = QLabel(f"welcome {appstate.getUser().fullname}")
        title.setStyleSheet("""
            font-size: 48px;
            color: white;
            font-weight: bold;
            margin-bottom: 20px;
        """)

        subtitle = QLabel("Glad to see you, excited to explore our new way of learning together. Let's dive in!")
        subtitle.setStyleSheet("""
            color: #94A3B8;
            font-size: 18px;
            margin-bottom: 30px;
        """)

        center_layout.addWidget(title, alignment=Qt.AlignCenter)
        center_layout.addWidget(subtitle, alignment=Qt.AlignCenter)
        center_layout.setAlignment(Qt.AlignCenter)

        # Add center content to header layout with stretch
        header_layout.addWidget(center_content, 1)

        # Add empty widget for right side balance (matches dropdown width)
        spacer = QWidget()
        spacer.setFixedSize(70, 65)
        header_layout.addWidget(spacer)

        main_layout.addWidget(header_container)

        # Create horizontal scroll area
        self.scroll_widget = HorizontalScrollArea()
        main_layout.addWidget(self.scroll_widget)

        courses = Subject.get_all_courses()


        
        for i, course in enumerate(courses):
            card = self.create_course_card(
                course["title"], 
                course["description"],
                course["chapters"],
                course["is_new"]
            )
            card.setFixedWidth(400)
            self.scroll_widget.add_widget(card)

        # Theme toggle and other settings
        self.theme_toggle = ThemeToggleButton(self)
        self.appstate.setThemeToggle(self.theme_toggle)
        self.theme_toggle.move(self.width() - 120, 20)

        self.setup_themes()
        self.theme_toggle.clicked.connect(lambda: self.apply_theme(self.theme_toggle.isChecked()))
        self.apply_theme(False)

        self.shortcut = QShortcut(QKeySequence('Esc'), self)
        self.shortcut.activated.connect(self.close)

        self.setWindowTitle("Interactive Quiz Platform")

    
   def resizeEvent(self, event):
    super().resizeEvent(event)
    self.theme_toggle.move(self.width() - 120, 20)
    margin = int(min(self.width(), self.height()) * 0.1)
    self.central_widget.layout().setContentsMargins(20, 20, margin, margin)
    
    
if __name__ == '__main__':
   app = QApplication(sys.argv)
   font = app.font()
   font.setFamily("Segoe UI")
   app.setFont(font)
   window = MCQHomePage()
   sys.exit(app.exec_())